import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def write_section(ws, start_row, title, headers, rows):
    ws.append([])
    ws.append([title])
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    return ws.max_row + 1

def main():
    with open("out/literacy_data_test.json", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Literacy Data"

    row = 1
    for student in data:
        # Section 1: Info
        info_keys = ["studentName", "studentId", "grade", "scaffolds", "district", "school", "generatedOn"]
        ws.append([f"Student Info"])
        ws.append(info_keys)
        ws.append([student.get(k, "") for k in info_keys])
        row = ws.max_row + 1

        ws.append([])  # 1 empty row

        # Section 2: Reading Activities
        ws.append(["Reading Activities"])
        reading = student.get("reading_activities", {})
        reading_data = reading.get("data", [])
        if reading_data:
            reading_headers = list(reading_data[0].keys())
            ws.append(reading_headers)
            for r in reading_data:
                ws.append([r.get(h, "") for h in reading_headers])
        else:
            ws.append(["No data"])
        row = ws.max_row + 1

        ws.append([])  # 1 empty row

        # Section 3: Open-ended Writing Activities
        ws.append(["Open-ended Writing Activities"])
        open_ended = student.get("open_ended_writing_activities", {})
        open_ended_data = open_ended.get("data", [])
        if open_ended_data:
            open_ended_headers = list(open_ended_data[0].keys())
            ws.append(open_ended_headers)
            for r in open_ended_data:
                ws.append([r.get(h, "") for h in open_ended_headers])
        else:
            ws.append(["No data"])
        row = ws.max_row + 1

        ws.append([])  # 1 empty row

        # Section 4: Writing Assignments
        ws.append(["Writing Assignments"])
        writing = student.get("writing_assignments", {})
        writing_data = writing.get("data", [])
        if writing_data:
            writing_headers = list(writing_data[0].keys())
            ws.append(writing_headers)
            for r in writing_data:
                ws.append([r.get(h, "") for h in writing_headers])
        else:
            ws.append(["No data"])
        row = ws.max_row + 2  # 2 empty rows between students
        ws.append([])
        ws.append([])

    # Optional: Auto-fit columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 50))

    wb.save("literacy_data_test.xlsx")
    print("Exported to literacy_data_test.xlsx")

if __name__ == "__main__":
    main()